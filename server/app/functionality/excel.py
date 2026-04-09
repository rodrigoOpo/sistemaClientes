from openpyxl import Workbook, load_workbook


#La ruta al archivo cambia dependiendo desde donde ejecutes el servidor.
#Si lo ejecutas desde la carpeta server entonces esta ruta sirve
#Lo importante es que no falte o sobre ninguna carpeta en la ruta al juntarse con la del servidor
wb = load_workbook('app/functionality/data/clientes.xlsx')
sheet = wb.active

#Iterar a través de filas
for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
    print(row)


#El método append() faciita la inserción de nuevas filas

#Las filas para añadir en el excel
client_data = [
    [5, "ROberto", "robert@noseque.con", "76453234"],
    [6, "Rupeto", "rupert@noseque.con", "76543234"],
    [7, "Roserto", "rosert@noseque.con", "764877334"],
]

for row in client_data:
    sheet.append(row)
wb.save("app/functionality/data/clientes.xlsx")

print("----------------------------------------------------------------")

for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
    print(row)

print("------------------------------------------------------------------")

#Iterar a través de las columnas
for column in sheet.iter_cols(min_col=1, max_col=4, values_only=True):
    print(column)

#Acceder a un rango de celdas
cell_range = sheet["A1:B3"]
for row in cell_range:
    for cell in row:
        print(cell.value)

#flujo de datos:
# formulario en el que el usuario pasa datos al endpoint.
# endpoint llama a la funcion 